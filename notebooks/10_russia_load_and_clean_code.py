# ===== CELL 1 =====import pandas as pd
from pathlib import Path

RAW_DIR = Path('../data/raw/russia')
OUT_DIR = Path('../data/processed')
OUT_DIR.mkdir(parents=True, exist_ok=True)

FILES = [
    'СПАРК_2012-2015.xlsx',
    'СПАРК_2016-2020.xlsx',
    'СПАРК_2021-2025.xlsx',
]

COLS_TO_DROP = [
    '№',
    'Наименование',
    'Мои списки',
    'Сумма незавершенных исков в роли ответчика, RUB',
    'Исполнительные производства, RUB',
]

ID_COL = 'Регистрационный номер'
# ===== CELL 2 =====def load_file_to_long(path: Path) -> pd.DataFrame:
    """
    Читает один xlsx-файл и сразу переводит в long-формат:
    строки = (company_id, year, metric, value).
    """
    df = pd.read_excel(path, sheet_name='report', header=3)

    # Убираем ненужные колонки
    drop = [c for c in COLS_TO_DROP if c in df.columns]
    df = df.drop(columns=drop)

    # Оставляем только ID + колонки с годами
    year_cols = [
        c for c in df.columns
        if c != ID_COL and any(str(y) in c for y in range(2010, 2030))
    ]
    df = df[[ID_COL] + year_cols].drop_duplicates(subset=ID_COL).dropna(subset=[ID_COL])

    # ID → int (убираем float / научную нотацию)
    df[ID_COL] = df[ID_COL].astype(int)

    # Melt: одна строка = (company_id, raw_col, value)
    melted = df.melt(
        id_vars=ID_COL,
        value_vars=year_cols,
        var_name='raw_col',
        value_name='value'
    )

    # Извлекаем год и название показателя
    melted['year'] = melted['raw_col'].str.extract(r'^(\d{4})').astype(int)
    melted['metric'] = (
        melted['raw_col']
        .str.replace(r'^\d{4},\s*', '', regex=True)
        .str.replace(r',\s*RUB$', '', regex=True)
        .str.strip()
    )

    return melted.drop(columns='raw_col')


def load_category(category: str) -> pd.DataFrame:
    """
    Загружает все файлы одной категории, объединяет в long-формат,
    убирает дубли, возвращает wide-формат: строки = (company_id, year).
    """
    parts = []
    for fname in FILES:
        path = RAW_DIR / category / fname
        long = load_file_to_long(path)
        years = sorted(long['year'].unique())
        n = long[ID_COL].nunique()
        print(f'  {fname}: {n} компаний, годы={years}')
        parts.append(long)

    combined = pd.concat(parts, ignore_index=True)

    # Убираем дубли: для каждой (company, year, metric) оставляем
    # первое непустое значение
    key = [ID_COL, 'year', 'metric']
    not_null = combined.dropna(subset=['value'])
    null_only = combined[combined['value'].isna()]
    combined = (
        pd.concat([not_null, null_only], ignore_index=True)
        .drop_duplicates(subset=key, keep='first')
    )

    # Pivot → wide: (company_id, year) × metrics
    wide = combined.pivot_table(
        index=[ID_COL, 'year'],
        columns='metric',
        values='value',
        aggfunc='first'
    ).reset_index()
    wide.columns.name = None

    print(f'  → итого: {wide[ID_COL].nunique()} компаний, '
          f'лет={sorted(wide["year"].unique())}, shape={wide.shape}')
    return wide
# ===== CELL 3 =====print('=== Active ===')
active = load_category('active')
active['is_bankrupt'] = 0

print()
print('=== Bankrupt ===')
bankrupt = load_category('bankrupt')
bankrupt['is_bankrupt'] = 1
# ===== CELL 4 =====# --- Очистка «Среднесписочная численность работников» ---
# В исходных данных: диапазоны ("0 - 5", "51 - 100"), числа с пробелами ("1 011"),
# одиночное значение "> 5 000", чистые числа. Приводим всё к float.

EMP_COL = 'Среднесписочная численность работников'

def parse_employee_count(val):
    """Парсит численность работников из текста в число."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    # "> 5 000" → 5000 (нижняя граница)
    if s.startswith('>'):
        return float(s[1:].replace(' ', ''))
    # "0 - 5" → 2.5  (среднее диапазона)
    if ' - ' in s:
        parts = s.split(' - ')
        try:
            lo = float(parts[0].replace(' ', ''))
            hi = float(parts[1].replace(' ', ''))
            return (lo + hi) / 2
        except ValueError:
            return None
    # "1 011" → 1011  (число с пробелами-разделителями)
    try:
        return float(s.replace(' ', ''))
    except ValueError:
        return None

for frame in [active, bankrupt]:
    frame[EMP_COL] = frame[EMP_COL].apply(parse_employee_count)

print(f'Active  — {EMP_COL}: dtype={active[EMP_COL].dtype}, NaN={active[EMP_COL].isna().sum()}')
print(f'Bankrupt — {EMP_COL}: dtype={bankrupt[EMP_COL].dtype}, NaN={bankrupt[EMP_COL].isna().sum()}')
print()
print('Примеры:', active[EMP_COL].dropna().unique()[:15])
# ===== CELL 5 =====# TARGET_WINDOW_V2 (K=2)
# Event-based target with K-year horizon window:
# is_bankrupt=1 on the last K years of each bankrupt company (the distressed window).
# Previous years of the same company remain class=0 (they were healthy then).
K_HORIZON = 2

# Объединяем active + bankrupt
df = pd.concat([active, bankrupt], ignore_index=True)

# --- Метка компании-банкрота (entity-level, для group-split и визуализации) ---
df['bankrupt_company'] = df['is_bankrupt'].copy()

# --- Target: is_bankrupt = 1 в последние K лет жизни компании-банкрота ---
fin_cols = [c for c in df.columns if c not in [ID_COL, 'year', 'is_bankrupt', 'bankrupt_company']]

# Последний год с данными у каждого банкрота
bankrupt_rows = df[df['bankrupt_company'] == 1].copy()
has_data = bankrupt_rows[fin_cols].notna().any(axis=1)
last_year_per_company = (
    bankrupt_rows.loc[has_data]
    .groupby(ID_COL)['year'].max()
    .reset_index()
    .rename(columns={'year': '_last_year'})
)

df = df.merge(last_year_per_company, on=ID_COL, how='left')
df['is_bankrupt'] = (
    (df['bankrupt_company'] == 1)
    & (df['year'] >= df['_last_year'] - (K_HORIZON - 1))
    & (df['year'] <= df['_last_year'])
).astype(int)
df = df.drop(columns='_last_year')

# --- Сортировка: банкроты первыми, потом действующие ---
df = df.sort_values(
    ['bankrupt_company', ID_COL, 'year'],
    ascending=[False, True, True]
).reset_index(drop=True)

service = ['bankrupt_company', 'is_bankrupt']
cols = [c for c in df.columns if c not in service] + service
df = df[cols]

print(f'Итоговый датасет: {df.shape}')
print(f'Годы: {sorted(df["year"].unique())}')
print(f'Активных компаний:  {df[df.bankrupt_company==0][ID_COL].nunique()}')
print(f'Компаний-банкротов: {df[df.bankrupt_company==1][ID_COL].nunique()}')
print(f'Строк с target=1 (окно K={K_HORIZON}): {(df.is_bankrupt==1).sum()}')
print(f'\nПоказатели ({len(fin_cols)}):')
print(fin_cols)

# ===== CELL 6 =====print('Пропуски по колонкам (%):')
print((df.isnull().mean() * 100).round(1).to_string())
# ===== CELL 7 =====# Первые строки — банкроты (видно bankrupt_company=1, is_bankrupt=1 только на последнем году)
print('=== Банкроты (первая компания) ===')
first_bankrupt_id = df.loc[df.bankrupt_company == 1, ID_COL].iloc[0]
display(df[df[ID_COL] == first_bankrupt_id][[ID_COL, 'year', 'bankrupt_company', 'is_bankrupt']].to_string(index=False))
print()
df.head(20)
# ===== CELL 8 =====# Сохраняем
# utf-8-sig — корректно открывается в Excel (с BOM)
out_path = OUT_DIR / 'russia_ml.csv'
df.to_csv(out_path, index=False, encoding='utf-8-sig')

print(f'Сохранено: {out_path}')
print(f'Размер файла: {out_path.stat().st_size / 1024 / 1024:.1f} MB')
# ===== CELL 9 =====from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

xlsx_path = OUT_DIR / 'russia_ml.xlsx'

# Пишем данные через pandas, потом форматируем через openpyxl
with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='data')

wb = load_workbook(xlsx_path)
ws = wb['data']

# --- Цвета ---
HEADER_FILL_ACTIVE   = PatternFill('solid', fgColor='1F4E79')   # тёмно-синий — ID/год/таргет
HEADER_FILL_METRICS  = PatternFill('solid', fgColor='2E75B6')   # синий — финансовые показатели
ROW_FILL_BANKRUPT    = PatternFill('solid', fgColor='FCE4D6')   # светло-оранжевый — банкрот
ROW_FILL_ACTIVE      = PatternFill('solid', fgColor='FFFFFF')   # белый — активная
ROW_FILL_TARGET      = PatternFill('solid', fgColor='F4B084')   # тёмно-оранжевый — строка с target=1

HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
DEFAULT_FONT = Font(size=10)
THIN_BORDER  = Border(
    bottom=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
)

# Индексы колонок (1-based)
col_names = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
service_cols = {ID_COL, 'year', 'is_bankrupt', 'bankrupt_company'}
bankrupt_company_col = col_names.index('bankrupt_company') + 1
is_bankrupt_col = col_names.index('is_bankrupt') + 1

# --- Шапка ---
for col_idx, col_name in enumerate(col_names, start=1):
    cell = ws.cell(1, col_idx)
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = HEADER_FILL_ACTIVE if col_name in service_cols else HEADER_FILL_METRICS

# --- Строки данных ---
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    is_bankrupt_co = row[bankrupt_company_col - 1].value
    is_target_row  = row[is_bankrupt_col - 1].value
    if is_target_row == 1:
        fill = ROW_FILL_TARGET       # последний год банкрота — тёмно-оранжевый
    elif is_bankrupt_co == 1:
        fill = ROW_FILL_BANKRUPT     # остальные годы банкрота — светло-оранжевый
    else:
        fill = ROW_FILL_ACTIVE       # действующая — белый
    for cell in row:
        cell.font = DEFAULT_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right' if isinstance(cell.value, (int, float)) else 'left')

# --- Ширина колонок ---
for col_idx, col_name in enumerate(col_names, start=1):
    letter = get_column_letter(col_idx)
    if col_name == ID_COL:
        ws.column_dimensions[letter].width = 22
    elif col_name in ('year', 'is_bankrupt', 'bankrupt_company'):
        ws.column_dimensions[letter].width = 13
    else:
        ws.column_dimensions[letter].width = 20

# --- Высота шапки и заморозка ---
ws.row_dimensions[1].height = 40
ws.freeze_panes = 'C2'

wb.save(xlsx_path)
print(f'Сохранено: {xlsx_path} ({xlsx_path.stat().st_size / 1024 / 1024:.1f} MB)')
print('Строки банкротов — оранжевые, строка target=1 — тёмно-оранжевая.')
# ===== CELL 11 =====import pandas as pd
from pathlib import Path

PROCESSED = Path('../data/processed')
df = pd.read_csv(PROCESSED / 'russia_ml.csv', encoding='utf-8-sig')

ID_COL = 'Регистрационный номер'
SERVICE_COLS = [ID_COL, 'year', 'is_bankrupt', 'bankrupt_company']
FIN_COLS = [c for c in df.columns if c not in SERVICE_COLS]

CORE_COLS = [
    'Активы  всего',
    'Выручка',
    'Капитал и резервы',
    'Оборотные активы',
    'Краткосрочные обязательства',
    'Чистая прибыль (убыток)',
]
SPARSE_ZERO_COLS = [
    'Нематериальные активы',
    'Долгосрочные обязательства',
    'Проценты к уплате',
    'Сальдо денежных потоков от текущих операций',
    'Внеоборотные активы',
]
RECOVERABLE_COLS = [c for c in FIN_COLS if c not in CORE_COLS + SPARSE_ZERO_COLS]

print(f'Всего строк на входе: {len(df):,}')
print(f'CORE ({len(CORE_COLS)}): {CORE_COLS}')
print(f'SPARSE_ZERO ({len(SPARSE_ZERO_COLS)}): {SPARSE_ZERO_COLS}')
print(f'RECOVERABLE ({len(RECOVERABLE_COLS)}): {RECOVERABLE_COLS}')

# ===== CELL 12 =====# --- Шаг 1: внутри каждой компании ffill + bfill по финансовым колонкам ---
df = df.sort_values([ID_COL, 'year']).reset_index(drop=True)
df[FIN_COLS] = (
    df.groupby(ID_COL, group_keys=False)[FIN_COLS]
      .apply(lambda g: g.ffill().bfill())
)
print('После ffill+bfill по компаниям:')
print((df[FIN_COLS].isna().mean() * 100).round(1).to_string())

# ===== CELL 13 =====# --- Шаг 2: разделяем активных и банкротов ---
active = df[df['bankrupt_company'] == 0].copy()
bankrupt = df[df['bankrupt_company'] == 1].copy()

shape_before = df.shape
print(f'Before cleaning: {shape_before}')
print(f'  активные строки:  {len(active):,}')
print(f'  банкротные строки: {len(bankrupt):,}')

# --- Активные: выкидываем строки без CORE-отчётности ---
active_core_missing = active[CORE_COLS].isna().any(axis=1)
print(f'\nАктивных строк без CORE-отчёта: {active_core_missing.sum():,} — drop')
active = active.loc[~active_core_missing].copy()

# --- Активные: SPARSE-ZERO → 0 (в СПАРКе пусто = не отражено) ---
active[SPARSE_ZERO_COLS] = active[SPARSE_ZERO_COLS].fillna(0)
# теоретически RECOVERABLE уже без NaN (после ffill+bfill внутри компании);
# если остатки есть — это изолированная компания с 1 годом и пустой ячейкой → 0
active[RECOVERABLE_COLS] = active[RECOVERABLE_COLS].fillna(0)

# --- Банкроты: ничего не выкидываем, остатки → 0 ---
bankrupt[FIN_COLS] = bankrupt[FIN_COLS].fillna(0)

df_clean = pd.concat([active, bankrupt], ignore_index=True)
df_clean = df_clean.sort_values(
    ['bankrupt_company', ID_COL, 'year'], ascending=[False, True, True]
).reset_index(drop=True)

print(f'\nAfter cleaning:  {df_clean.shape}')
print(f'Строк удалено: {shape_before[0] - df_clean.shape[0]:,}')
print(f'NaN в финансовых колонках после чистки: {df_clean[FIN_COLS].isna().sum().sum()}')

# ===== CELL 14 =====# --- Контроль: состав классов и уникальных компаний ---
n_companies_total = df_clean[ID_COL].nunique()
n_companies_bankrupt = df_clean.loc[df_clean.bankrupt_company == 1, ID_COL].nunique()
n_companies_active = df_clean.loc[df_clean.bankrupt_company == 0, ID_COL].nunique()

print('=== Уникальные компании ===')
print(f'  активных:  {n_companies_active:,}')
print(f'  банкротов: {n_companies_bankrupt:,}')
print(f'  всего:     {n_companies_total:,}')

print('\n=== Target (is_bankrupt) на уровне строк-лет ===')
print(df_clean['is_bankrupt'].value_counts().rename({0: 'Target=0', 1: 'Target=1'}).to_string())
ratio = df_clean['is_bankrupt'].value_counts(normalize=True)[1]
print(f'\nДоля позитивного класса: {ratio*100:.3f}%  '
      f'(imbalance ≈ 1:{int((1-ratio)/ratio)})')
print('Рекомендуемый class_weight ≈ {0: 1, 1: 10} (задаётся в моделях, не в данных).')

# ===== CELL 15 =====# --- Сохраняем финальный датасет ---
OUT = PROCESSED / 'ru_panel_cleaned.csv'
df_clean.to_csv(OUT, index=False, encoding='utf-8-sig')
print(f'✅ Сохранено: {OUT}  ({OUT.stat().st_size/1024/1024:.1f} MB)')
print(f'Shape: {df_clean.shape}')
